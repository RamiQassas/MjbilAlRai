# MjbilAlRai_App/admin.py

from django.contrib import admin
from django.http import HttpResponse
from .models import Reservation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class ReservationAdmin(admin.ModelAdmin):
    # الحقول التي سيتم عرضها في صفحة إدارة Django Admin
    list_display = ('reservation_number', 'customer_name', 'carpenter_name', 'reservation_date', 'is_approved', 'is_rejected', 'is_completed', 'status')

    # إضافة خيارات الفلترة على البيانات
    list_filter = ('is_approved', 'is_rejected', 'is_completed', 'reservation_date', 'status')

    # إضافة البحث باستخدام الحقول
    search_fields = ('customer_name', 'reservation_number', 'carpenter_name')

    # إضافة الإجراءات المتاحة مثل التصدير كملف Excel
    actions = ['export_as_excel']

    # دالة مخصصة لتصدير البيانات كملف Excel
    def export_as_excel(self, request, queryset):
        # إنشاء ملف إكسل جديد
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'الحجوزات'

        headers = [
            'رقم الحجز', 'اسم العميل', 'اسم النجار', 'نوع الخرسانة',
            'كمية الخرسانة', 'موقع الصب', 'المسافة التقديرية',
            'تاريخ الحجز', 'تاريخ الموافقة', 'رسالة الموافقة',
            'السعر للوحدة', 'الخصم', 'إجمالي التكلفة', 'ملاحظات المحاسب',
            'مجموع الدفعات المدفوعة', 'المبلغ المتبقي', 'اكتمال الحجز', 'تاريخ اكتمال الحجز', 'الحالة'
        ]
        worksheet.append(headers)

        # تنسيق العناوين
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill

        # إعداد حدود الخلايا
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # كتابة بيانات الحجوزات
        for obj in queryset:
            row = [
                obj.reservation_number,
                obj.customer_name,
                obj.carpenter_name,
                obj.concrete_type,
                obj.concrete_quantity,
                obj.site_location,
                obj.estimated_distance,
                obj.reservation_date.strftime('%Y-%m-%d') if obj.reservation_date else '',
                obj.approval_date.strftime('%Y-%m-%d') if obj.approval_date else '',
                obj.approval_message,
                obj.price_per_unit,
                obj.discount,
                obj.total_cost,
                obj.accountant_notes,
                obj.payments,
                obj.remaining_balance,
                'نعم' if obj.is_completed else 'لا',
                obj.completion_date.strftime('%Y-%m-%d') if obj.completion_date else '',
                obj.status,  # إضافة الحقل الجديد للحالة
            ]
            worksheet.append(row)

        # تنسيق الصفوف
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=19):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

        # إعداد الاستجابة وإرسال الملف للمستخدم
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=الحجوزات.xlsx'
        workbook.save(response)
        return response

    # تحديد الوصف لزر الإجراء في صفحة الإدارة
    export_as_excel.short_description = "تصدير مختار كملف Excel"

# تسجيل النموذج داخل صفحة Django Admin مع الخيارات المخصصة
admin.site.register(Reservation, ReservationAdmin)
