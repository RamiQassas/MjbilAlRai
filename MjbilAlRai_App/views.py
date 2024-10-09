# MjbilAlRai_App/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth import login, authenticate, logout
from django.contrib import messages
from django.contrib.auth.forms import AuthenticationForm
from django.http import HttpResponse
from django.db.models import Q, Sum
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from decimal import Decimal
from datetime import date
from .forms import ReservationForm, FinancialDetailsForm, PaymentForm
from .models import Reservation
import logging

# إعداد السجلات (Logging)
logger = logging.getLogger(__name__)

# رسائل ثابتة للمستخدم
SUCCESS_LOGIN_MSG = 'لقد تم الدخول بنجاح!'
ERROR_LOGIN_MSG = 'اسم المستخدم أو كلمة المرور غير صحيحة.'

# ديكورات مخصصة للصلاحيات
def has_permission(user, perm):
    return user.has_perm(perm)

manage_permission_required = permission_required('MjbilAlRai_App.can_manage_reservations', raise_exception=True)
confirm_permission_required = permission_required('MjbilAlRai_App.can_confirm_reservations', raise_exception=True)

# دمج صلاحية تعديل وعرض محاسب إلى صلاحية واحدة
manage_accountant_permission_required = permission_required('MjbilAlRai_App.can_manage_accountant', raise_exception=True)

# Helper Functions
def calculate_total_cost(reservation):
    return (reservation.price_per_unit or Decimal(0)) * (reservation.concrete_quantity or Decimal(0))

def calculate_remaining_balance(reservation):
    total_cost = calculate_total_cost(reservation)
    discounted_cost = total_cost - (reservation.discount or Decimal(0))
    remaining = discounted_cost - (reservation.payments or Decimal(0))
    return remaining

# العروض (Views)

# الصفحة الرئيسية
# الصفحة الرئيسية
# الصفحة الرئيسية
def home(request):
    reservations = []
    total_gross_amount_sum = total_discount_sum = total_payments_sum = total_remaining_sum = 0
    error = None

    if request.method == 'POST':
        phone_number = request.POST.get('phone_number')
        reservation_number = request.POST.get('reservation_number')

        if phone_number:
            reservations = Reservation.objects.filter(phone_number=phone_number)
            if reservations.exists():
                total_gross_amount_sum = sum(calculate_total_cost(r) for r in reservations)
                total_discount_sum = sum(r.discount or Decimal(0) for r in reservations)
                total_payments_sum = reservations.aggregate(Sum('payments'))['payments__sum'] or Decimal(0)
                total_remaining_sum = sum(calculate_remaining_balance(r) for r in reservations)
            else:
                error = "لا توجد حجوزات مسجلة لهذا الرقم."
        elif reservation_number:
            reservation = Reservation.objects.filter(reservation_number=reservation_number).first()
            if reservation:
                return render(request, 'MjbilAlRai_App/reservation_status.html', {'reservation': reservation})
            else:
                error = "رقم الحجز غير موجود أو الحجز مكتمل."
        else:
            error = "يرجى إدخال رقم الهاتف أو رقم الحجز."

    context = {
        'reservations': reservations,
        'total_gross_amount_sum': total_gross_amount_sum,
        'total_discount_sum': total_discount_sum,
        'total_payments_sum': total_payments_sum,
        'total_remaining_sum': total_remaining_sum,
        'error': error
    }
    return render(request, 'MjbilAlRai_App/home.html', context)



# إدارة الحجوزات
# إدارة الحجوزات
@login_required
@manage_permission_required
def manage_reservations(request):
    status = request.GET.get('status')
    reservation_date = request.GET.get('reservation_date')

    # إنشاء فلتر ديناميكي
    filters = Q()
    if status:
        if status == 'معلق':
            filters &= Q(is_approved=False, is_rejected=False, is_confirmed=False)
        elif status == 'مقبول':
            filters &= Q(is_approved=True, is_confirmed=False)
        elif status == 'مرفوض':
            filters &= Q(is_rejected=True)
        elif status == 'مكتمل':
            filters &= Q(is_completed=True)

    # فلترة حسب تاريخ الحجز إذا تم إدخال التاريخ
    if reservation_date:
        filters &= Q(reservation_date=reservation_date)

    # تطبيق الفلاتر على الحجوزات
    reservations = Reservation.objects.filter(filters)

    # عرض الصفحة مع الحجوزات المفلترة
    return render(request, 'MjbilAlRai_App/manage_reservations.html', {'reservations': reservations})


# تأكيد الحجوزات
@login_required
@confirm_permission_required
def confirm_reservations(request):
    if request.method == 'POST':
        reservation_id = request.POST.get('reservation_id')
        concrete_quantity = request.POST.get('concrete_quantity')
        concrete_type = request.POST.get('concrete_type')
        is_completed = request.POST.get('is_completed') == 'true'

        reservation = get_object_or_404(Reservation, id=reservation_id)

        try:
            concrete_quantity = Decimal(concrete_quantity)

            reservation.concrete_quantity = concrete_quantity
            reservation.concrete_type = concrete_type
            reservation.is_confirmed = True

            # تحديث حالة الحجز بناءً على التأكيد
            if is_completed:
                reservation.is_completed = True
                reservation.status = 'مكتمل'
                reservation.completion_date = date.today()
            else:
                reservation.is_completed = False
                reservation.status = 'مقبول'
                reservation.completion_date = None

            reservation.save()

            messages.success(request, f"تم تأكيد الحجز رقم {reservation.reservation_number} بنجاح.")
        except Exception as e:
            messages.error(request, f"حدث خطأ أثناء تأكيد الحجز: {e}")

        return redirect('confirm_reservations')

    # عرض الحجوزات التي تحتاج إلى تأكيد فقط (مقبولة ولكن غير مكتملة)
    reservations = Reservation.objects.filter(is_approved=True, is_rejected=False, is_confirmed=False)
    return render(request, 'MjbilAlRai_App/confirm_reservations.html', {
        'reservations': reservations,
    })

# قبول الحجز
@login_required
@manage_permission_required
def approve_reservation(request, reservation_id):
    reservation = get_object_or_404(Reservation, id=reservation_id)
    if request.method == 'POST':
        approval_date = request.POST.get('approval_date')
        approval_message = request.POST.get('approval_message', '')

        reservation.is_approved = True
        reservation.is_rejected = False
        reservation.approval_date = approval_date
        reservation.approval_message = approval_message
        reservation.save()

        # تحديث حالة الحجز
        reservation.status = 'مقبول'
        reservation.save()

        messages.success(request, f"تم قبول الحجز رقم {reservation.reservation_number} بنجاح.")
    return redirect('manage_reservations')

# رفض الحجز
@login_required
@manage_permission_required
def reject_reservation(request, reservation_id):
    reservation = get_object_or_404(Reservation, id=reservation_id)
    reservation.is_approved = False
    reservation.is_rejected = True
    reservation.status = 'مرفوض'
    reservation.save()
    messages.success(request, f"تم رفض الحجز رقم {reservation.reservation_number}.")
    return redirect('manage_reservations')

# تسجيل الدخول
def login_user(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, SUCCESS_LOGIN_MSG)

                # توجيه المستخدم بناءً على الصلاحيات
                if has_permission(user, 'MjbilAlRai_App.can_manage_accountant'):
                    return redirect('accountant_dashboard')
                elif has_permission(user, 'MjbilAlRai_App.can_manage_reservations'):
                    return redirect('manage_reservations')
                elif has_permission(user, 'MjbilAlRai_App.can_confirm_reservations'):
                    return redirect('confirm_reservations')
                else:
                    return redirect('home')
            else:
                messages.error(request, ERROR_LOGIN_MSG)
        else:
            messages.error(request, ERROR_LOGIN_MSG)
    else:
        form = AuthenticationForm()
    return render(request, 'MjbilAlRai_App/login.html', {'form': form})

# تسجيل الخروج
@login_required
def logout_user(request):
    logout(request)
    messages.success(request, 'تم تسجيل الخروج بنجاح!')
    return redirect('home')

# إضافة حجز جديد
def new_reservation(request):
    if request.method == 'POST':
        form = ReservationForm(request.POST)
        if form.is_valid():
            reservation = form.save()
            return render(request, 'MjbilAlRai_App/reservation_success.html', {'reservation': reservation})
        else:
            messages.error(request, "حدث خطأ أثناء إدخال الحجز. يرجى التحقق من البيانات المدخلة.")
    else:
        form = ReservationForm()
    return render(request, 'MjbilAlRai_App/new_reservation.html', {'form': form})

# تصدير الحجوزات إلى ملف Excel (عام)
@login_required
@manage_permission_required
def export_reservations(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "الحجوزات"

    headers = [
        'رقم الحجز', 'اسم العميل', 'اسم النجار', 'نوع الخرسانة',
        'كمية الخرسانة', 'موقع الصب', 'المسافة التقديرية',
        'تاريخ الحجز', 'تاريخ الموافقة', 'رسالة الموافقة',
        'السعر للوحدة', 'الخصم', 'إجمالي التكلفة', 'ملاحظات المحاسب',
        'مجموع الدفعات المدفوعة', 'المبلغ المتبقي', 'اكتمال الحجز', 'تاريخ اكتمال الحجز', 'الحالة'
    ]
    worksheet.append(headers)

    reservations = Reservation.objects.all()
    total_payments = Decimal(0)
    total_remaining = Decimal(0)

    for reservation in reservations:
        total_cost = calculate_total_cost(reservation)
        remaining_balance = calculate_remaining_balance(reservation)

        worksheet.append([
            reservation.reservation_number,
            reservation.customer_name,
            reservation.carpenter_name,
            reservation.concrete_type,
            reservation.concrete_quantity,
            reservation.site_location,
            reservation.estimated_distance,
            reservation.reservation_date.strftime('%Y-%m-%d') if reservation.reservation_date else '',
            reservation.approval_date.strftime('%Y-%m-%d') if reservation.approval_date else '',
            reservation.approval_message,
            reservation.price_per_unit,
            reservation.discount,
            reservation.total_cost,
            reservation.accountant_notes,
            reservation.payments,
            remaining_balance,
            'نعم' if reservation.is_completed else 'لا',
            reservation.completion_date.strftime('%Y-%m-%d') if reservation.completion_date else '',
            reservation.status,  # إضافة الحقل الجديد للحالة
        ])

        total_payments += reservation.payments or Decimal(0)
        total_remaining += remaining_balance

    # إضافة صف الإجماليات في نهاية الجدول
    worksheet.append([])
    worksheet.append(['', '', '', '', '', '', '', '', '', '', '', '', 'إجمالي:', '', total_payments, total_remaining, '', '', ''])

    # تنسيق الصفوف
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=19):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border

    # إعداد الاستجابة وإرسال الملف للمستخدم
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reservations.xlsx"'

    workbook.save(response)
    return response

# تصدير حجوزات العملاء إلى ملف Excel
def export_customer_reservations(request):
    if request.method == 'POST':
        phone_number = request.POST.get('phone_number')
        logger.debug(f"Exporting reservations for phone number: {phone_number}")
        if phone_number:
            reservations = Reservation.objects.filter(phone_number=phone_number)
            if reservations.exists():
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "حجوزات العميل"

                # إعداد نمط العناوين
                header_font = Font(bold=True, color="FFFFFF")
                header_alignment = Alignment(horizontal='center', vertical='center')
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

                # إعداد العناوين للجدول
                headers = [
                    'رقم الحجز', 'اسم العميل', 'اسم النجار', 'عيار الخرسانة',
                    'كمية الخرسانة', 'السعر للوحدة', 'الخصم', 'إجمالي التكلفة',
                    'ملاحظات المحاسب', 'مجموع الدفعات المدفوعة', 'المبلغ المتبقي',
                    'اكتمال الحجز', 'تاريخ اكتمال الحجز', 'الحالة'
                ]
                worksheet.append(headers)

                # تنسيق العناوين
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.fill = header_fill

                # إعداد حدود الخلايا
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # حساب البيانات وإضافتها للجدول
                total_remaining = Decimal(0)

                for idx, reservation in enumerate(reservations, start=2):
                    total_cost = calculate_total_cost(reservation)
                    remaining_balance = calculate_remaining_balance(reservation)

                    row_data = [
                        reservation.reservation_number,
                        reservation.customer_name,
                        reservation.carpenter_name,
                        reservation.concrete_type,
                        reservation.concrete_quantity,
                        reservation.price_per_unit,
                        reservation.discount,
                        reservation.total_cost,
                        reservation.accountant_notes,
                        reservation.payments,
                        remaining_balance,
                        'نعم' if reservation.is_completed else 'لا',
                        reservation.completion_date.strftime('%Y-%m-%d') if reservation.completion_date else '',
                        reservation.status,  # إضافة الحقل الجديد للحالة
                    ]
                    worksheet.append(row_data)

                    # تنسيق الصفوف
                    for cell in worksheet[idx]:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border

                    total_remaining += remaining_balance

                # إضافة صف الإجماليات في نهاية الجدول
                worksheet.append([])
                worksheet.append(['', '', '', '', '', '', '', 'إجمالي:', '', '', total_remaining, '', '', ''])

                # تنسيق صف الإجماليات
                total_row_idx = len(reservations) + 3
                for idx, cell in enumerate(worksheet[total_row_idx], start=1):
                    if idx == 8 or idx == 11:
                        cell.font = Font(bold=True, color="FF0000")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                        cell.border = thin_border

                # إعداد الاستجابة وإرسال الملف للمستخدم
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="customer_reservations_{phone_number}.xlsx"'

                workbook.save(response)
                return response
            else:
                messages.error(request, "لا توجد حجوزات مسجلة بهذا الرقم لتصديرها.")
                return redirect('customer_reservations')
        else:
            messages.error(request, "يرجى إدخال رقم الهاتف لتصدير الحجوزات.")
            return redirect('customer_reservations')
    else:
        messages.error(request, "طريقة الطلب غير صحيحة.")
        return redirect('customer_reservations')

# عرض حجوزات العملاء بناءً على رقم الهاتف
def customer_reservations(request):
    reservations = []
    total_gross_amount_sum = total_discount_sum = total_payments_sum = total_remaining_sum = 0
    error = None

    if request.method == 'POST':
        phone_number = request.POST.get('phone_number')

        if phone_number:
            reservations = Reservation.objects.filter(phone_number=phone_number)
            if reservations.exists():
                total_gross_amount_sum = sum(calculate_total_cost(r) for r in reservations)
                total_discount_sum = sum(r.discount or Decimal(0) for r in reservations)
                total_payments_sum = reservations.aggregate(Sum('payments'))['payments__sum'] or Decimal(0)
                total_remaining_sum = sum(calculate_remaining_balance(r) for r in reservations)
            else:
                error = "لا توجد حجوزات بهذا الرقم."
        else:
            error = "يرجى إدخال رقم الهاتف."

    context = {
        'reservations': reservations,
        'total_gross_amount_sum': total_gross_amount_sum,
        'total_discount_sum': total_discount_sum,
        'total_payments_sum': total_payments_sum,
        'total_remaining_sum': total_remaining_sum,
        'error': error
    }
    return render(request, 'MjbilAlRai_App/customer_reservations.html', context)

# لوحة المحاسب
@login_required
@login_required
@manage_accountant_permission_required
def accountant_dashboard(request):
    financial_status = request.GET.get('financial_status')

    # تصفية الحجوزات بناءً على الحالة المالية المحددة
    reservations = Reservation.objects.filter(is_confirmed=True)

    if financial_status == 'remaining':
        reservations = reservations.filter(remaining_balance__gt=0)
    elif financial_status == 'paid':
        reservations = reservations.filter(remaining_balance=0)
    elif financial_status == 'pending':
        reservations = reservations.filter(remaining_balance__lt=0)

    # حساب الإجماليات
    total_gross_amount_sum = sum(calculate_total_cost(r) for r in reservations)
    total_discount_sum = sum(r.discount or Decimal(0) for r in reservations)
    total_payments_sum = reservations.aggregate(Sum('payments'))['payments__sum'] or Decimal(0)
    total_remaining_sum = sum(calculate_remaining_balance(r) for r in reservations)

    # تمرير البيانات إلى القالب
    context = {
        'reservations': reservations,
        'total_gross_amount_sum': total_gross_amount_sum,
        'total_discount_sum': total_discount_sum,
        'total_payments_sum': total_payments_sum,
        'total_remaining_sum': total_remaining_sum,
        'financial_status': financial_status,
    }

    return render(request, 'MjbilAlRai_App/accountant_dashboard.html', context)



# تحديث التفاصيل المالية لحجز
# تحديث التفاصيل المالية لحجز
@login_required
@manage_accountant_permission_required  # تم استخدام الصلاحية المدمجة
def update_financial_details(request, reservation_id):
    reservation = get_object_or_404(Reservation, id=reservation_id)
    logger.debug(f"Processing update_financial_details for reservation ID: {reservation_id}")

    # حساب التكلفة الإجمالية عند التحميل الأول للصفحة (GET request)
    total_cost = calculate_total_cost(reservation)

    # حساب الرصيد المتبقي
    remaining_balance = calculate_remaining_balance(reservation)

    # حساب إجماليات الحجوزات المؤكدة
    reservations = Reservation.objects.filter(is_confirmed=True)
    total_gross_amount_sum = sum(calculate_total_cost(r) for r in reservations)
    total_discount_sum = sum(r.discount or Decimal(0) for r in reservations)
    total_payments_sum = reservations.aggregate(Sum('payments'))['payments__sum'] or Decimal(0)
    total_remaining_sum = sum(calculate_remaining_balance(r) for r in reservations)

    if request.method == 'POST':
        if 'update_financial' in request.POST:
            logger.debug("Received POST request for updating financial details.")
            financial_form = FinancialDetailsForm(request.POST, instance=reservation)
            if financial_form.is_valid():
                financial_form.save()
                logger.debug("Financial details updated successfully.")
                messages.success(request, f"تم تحديث البيانات المالية للحجز رقم {reservation.reservation_number} بنجاح.")
                return redirect('accountant_dashboard')
            else:
                logger.debug(f"Financial form errors: {financial_form.errors}")
                messages.error(request, "حدث خطأ أثناء تحديث البيانات المالية. يرجى التحقق من البيانات المدخلة.")
                payment_form = PaymentForm()
        elif 'record_payment' in request.POST:
            logger.debug("Received POST request for recording a payment.")
            payment_form = PaymentForm(request.POST, remaining_balance=reservation.remaining_balance or Decimal(0))
            if payment_form.is_valid():
                payment_amount = payment_form.cleaned_data.get('payment_amount') or Decimal(0)
                reservation.payments += payment_amount

                # إعادة حساب الرصيد المتبقي بعد تحديث المدفوعات
                remaining_balance = calculate_remaining_balance(reservation)
                reservation.remaining_balance = remaining_balance

                # إذا أصبح الرصيد المتبقي صفرًا أو أقل، نقوم بتعيين الحجز كمكتمل
                if reservation.remaining_balance <= 0 and not reservation.is_completed:
                    reservation.is_completed = True
                    reservation.status = 'مكتمل'  # تحديث الحالة إلى مكتمل
                    reservation.completion_date = date.today()

                reservation.save()
                logger.debug(f"Payment of {payment_amount} recorded successfully.")
                messages.success(request, f"تم تسجيل الدفعة بقيمة {payment_amount} دولار أمريكي للحجز رقم {reservation.reservation_number} بنجاح.")
                return redirect('accountant_dashboard')
            else:
                logger.debug(f"Payment form errors: {payment_form.errors}")
                messages.error(request, "حدث خطأ أثناء تسجيل الدفعة. يرجى التحقق من البيانات المدخلة.")
                financial_form = FinancialDetailsForm(instance=reservation)
        else:
            logger.debug("Received POST request with unknown action.")
            financial_form = FinancialDetailsForm(instance=reservation)
            payment_form = PaymentForm()
    else:
        financial_form = FinancialDetailsForm(instance=reservation)
        payment_form = PaymentForm()

    return render(request, 'MjbilAlRai_App/update_financial_details.html', {
        'financial_form': financial_form,
        'payment_form': payment_form,
        'reservation': reservation,
        'total_cost': total_cost,
        'remaining_balance': remaining_balance,
        'total_gross_amount_sum': total_gross_amount_sum,
        'total_discount_sum': total_discount_sum,
        'total_payments_sum': total_payments_sum,
        'total_remaining_sum': total_remaining_sum
    })

